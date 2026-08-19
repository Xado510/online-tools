import base64
import gzip
import json
import mimetypes
import os
import secrets
import socket
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from http.server import BaseHTTPRequestHandler, HTTPServer
from socketserver import ThreadingMixIn


class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True

from gmssl import sm2, sm4
from websocket import create_connection


ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
PRINT_SCRIPT = os.path.join(ROOT_DIR, "trigger_barcode_print.ps1")
EDGE_CANDIDATES = [
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
]
DEFAULT_BAR_PRINTER = (
    r"D:\we\data\xwechat_files\wxid_4m1kqcgnvn2v22_d9be\msg\file\2026-07"
    r"\条码打印软件\条码打印软件\BarPrinter.exe"
)
JD_BASE = "https://baozang-out.jd.com"
JD_SERVICE_BASE = "https://jdservice.jdl.com"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36"
LATEST_RESULTS = {}
JDL_TOKEN = ""
JDL_COOKIE = ""
LAST_BARCODE_ERROR = ""
DIGITAL_CONFIG = {"cookie": "", "userId": "", "appCode": "", "shopCode": ""}
DIGITAL_CONFIG_FILE = os.path.join(ROOT_DIR, "digital_config.json")
JDL_TOKEN_FILE = os.path.join(ROOT_DIR, "jdl_token.json")
CLIENT_CONFIGS = {}
CLIENT_JDL_TOKENS = {}
CLIENT_JDL_COOKIES = {}
SHARED_STATES = {}
SHARED_STATE_FILE = os.path.join(ROOT_DIR, "shared_state.json")
PERFORMANCE_MAP = {
    "01": "维修",
    "REPAIR": "维修",
    "02": "换新",
    "CHANGE_NEW": "换新",
    "05": "补贴",
    "SUBSIDY": "补贴",
    "08": "退货",
    "SALES_RETURN": "退货",
    "10": "增值服务",
    "ADD_VALUE": "增值服务",
    "11": "服务中台",
    "SERVICE_PLATFORM_THIRD_PART_RIGHT": "服务中台",
}


def decode_body(raw, headers):
    if headers.get("Content-Encoding", "").lower() == "gzip":
        return gzip.decompress(raw).decode("utf-8", "ignore")
    return raw.decode("utf-8", "ignore")


def load_digital_config():
    try:
        with open(DIGITAL_CONFIG_FILE, encoding="utf-8") as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            DIGITAL_CONFIG.update(
                {
                    key: str(data.get(key) or "").strip()
                    for key in ("cookie", "userId", "appCode", "shopCode")
                }
            )
    except Exception:
        pass


def save_digital_config():
    try:
        with open(DIGITAL_CONFIG_FILE, "w", encoding="utf-8") as handle:
            json.dump(DIGITAL_CONFIG, handle, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_jdl_token():
    global JDL_TOKEN, JDL_COOKIE
    try:
        with open(JDL_TOKEN_FILE, encoding="utf-8") as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            JDL_TOKEN = str(data.get("jdlToken") or "").strip()
            JDL_COOKIE = str(data.get("jdlCookie") or "").strip()
    except Exception:
        pass


def save_jdl_token():
    try:
        with open(JDL_TOKEN_FILE, "w", encoding="utf-8") as handle:
            json.dump(
                {"jdlToken": JDL_TOKEN, "jdlCookie": JDL_COOKIE},
                handle,
                ensure_ascii=False,
                indent=2,
            )
    except Exception:
        pass


def load_shared_states():
    try:
        with open(SHARED_STATE_FILE, encoding="utf-8") as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            SHARED_STATES.update(data)
    except Exception:
        pass


def save_shared_states():
    try:
        with open(SHARED_STATE_FILE, "w", encoding="utf-8") as handle:
            json.dump(SHARED_STATES, handle, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _state_timestamp(item):
    if not isinstance(item, dict):
        return ""
    return str(
        item.get("updatedAt")
        or item.get("unpackedAt")
        or item.get("createdAt")
        or ""
    )


def _merge_state_list(stored, incoming, id_key="id"):
    combined = {}
    for item in list(stored or []) + list(incoming or []):
        if not isinstance(item, dict) or not item.get(id_key):
            continue
        key = str(item[id_key])
        if key not in combined or _state_timestamp(item) >= _state_timestamp(combined[key]):
            combined[key] = item
    return list(combined.values())


def merge_shared_state(stored, incoming):
    stored = stored or {}
    incoming = incoming or {}
    return {
        "parcels": _merge_state_list(
            stored.get("parcels"), incoming.get("parcels")
        ),
        "anomalies": _merge_state_list(
            stored.get("anomalies"), incoming.get("anomalies")
        ),
    }


def find_barcode_printer():
    if os.path.isfile(DEFAULT_BAR_PRINTER):
        return DEFAULT_BAR_PRINTER
    roots = [
        r"D:\we\data\xwechat_files",
        r"C:\Users\zx173\Desktop",
        r"C:\Users\zx173\Downloads",
        r"C:\Users\zx173\Documents",
        "D:\\",
        r"C:\Program Files",
        r"C:\Program Files (x86)",
    ]
    skip_dirs = {
        "windows",
        "programdata",
        "$recycle.bin",
        "system volume information",
        "node_modules",
        ".git",
        "appdata",
    }
    for root in roots:
        if not os.path.isdir(root):
            continue
        for current, dirs, files in os.walk(root):
            depth = current[len(root):].count(os.sep)
            if depth >= 5:
                dirs[:] = []
            dirs[:] = [
                name
                for name in dirs
                if name.lower() not in skip_dirs
                and not name.startswith("$")
                and not name.startswith(".")
            ]
            try:
                for name in files:
                    if name.lower() == "barprinter.exe":
                        return os.path.join(current, name)
                for name in dirs:
                    if "条码打印软件" in name:
                        nested = os.path.join(current, name, "BarPrinter.exe")
                        if os.path.isfile(nested):
                            return nested
            except Exception:
                continue
    return ""


def find_key(obj, key):
    if isinstance(obj, dict):
        if key in obj and obj[key] not in (None, ""):
            return obj[key]
        for value in obj.values():
            result = find_key(value, key)
            if result not in (None, ""):
                return result
    elif isinstance(obj, list):
        for value in obj:
            result = find_key(value, key)
            if result not in (None, ""):
                return result
    return None


def extract_cookie_value(cookie_text, name):
    for part in (cookie_text or "").split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        if key.strip() == name:
            return value.strip()
    return ""


def call_jd(path, payload, cookie, user_id, app_code):
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "identity",
        "Content-Type": "application/json",
        "Origin": "https://digital-ins.jd.com",
        "Referer": "https://digital-ins.jd.com/repair",
        "sysType": "1",
        "User-Agent": UA,
    }
    if cookie:
        headers["Cookie"] = cookie.replace("\r", "").replace("\n", "")
    if user_id:
        headers["userId"] = user_id.strip()
    if app_code:
        headers["appcode"] = app_code.strip()
    request = urllib.request.Request(
        JD_BASE + path,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            body = decode_body(response.read(), response.headers)
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = decode_body(error.read(), error.headers)
        try:
            return json.loads(body)
        except Exception:
            return {"success": False, "error": f"HTTP {error.code}: {body[:200]}"}
    except Exception as error:
        return {"success": False, "error": str(error)}


def call_jd_get(path, cookie, user_id, app_code):
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "identity",
        "Content-Type": "application/json",
        "Origin": "https://digital-ins.jd.com",
        "Referer": "https://digital-ins.jd.com/repair",
        "sysType": "1",
        "User-Agent": UA,
    }
    if cookie:
        headers["Cookie"] = cookie.replace("\r", "").replace("\n", "")
    if user_id:
        headers["userId"] = user_id.strip()
    if app_code:
        headers["appcode"] = app_code.strip()
    request = urllib.request.Request(
        JD_BASE + path,
        headers=headers,
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            body = decode_body(response.read(), response.headers)
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = decode_body(error.read(), error.headers)
        try:
            return json.loads(body)
        except Exception:
            return {"success": False, "error": f"HTTP {error.code}: {body[:200]}"}
    except Exception as error:
        return {"success": False, "error": str(error)}


def jd_encrypt_data(key_text, payload):
    decoded = base64.b64decode(key_text)
    if len(decoded) < 65:
        raise RuntimeError("SM key too short")
    public_key = decoded[-65:].hex()
    public_key_header = decoded[:-65].hex()

    sm4_key = secrets.token_hex(8)
    iv = secrets.token_hex(16)

    crypt_sm2 = sm2.CryptSM2(private_key="", public_key=public_key, mode=0)
    sm2_encrypted = crypt_sm2.encrypt(sm4_key.encode("ascii")).hex()

    sm4_key_bytes = sm4_key.encode("ascii")
    crypt_sm4 = sm4.CryptSM4()
    crypt_sm4.set_key(sm4_key_bytes, sm4.SM4_ENCRYPT)
    plaintext = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    sm4_encrypted = crypt_sm4.crypt_cbc(bytes.fromhex(iv), plaintext).hex()

    combined = (
        bytes.fromhex(public_key_header)
        + bytes.fromhex(sm2_encrypted)
        + bytes.fromhex(iv)
        + bytes.fromhex(sm4_encrypted)
    )
    return base64.b64encode(combined).decode("ascii")


def _find_edge_path():
    for candidate in EDGE_CANDIDATES:
        if os.path.isfile(candidate):
            return candidate
    raise RuntimeError("Edge browser not found")


def _wait_debugger(port, timeout=20):
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            request = urllib.request.Request(
                "http://127.0.0.1:%d/json/version" % port,
                headers={"User-Agent": "Codex/1.0"},
            )
            with urllib.request.urlopen(request, timeout=2) as response:
                return json.loads(response.read().decode("utf-8", "ignore"))
        except Exception:
            time.sleep(0.3)
    raise RuntimeError("Edge debugger did not start")


def _get_page_ws(port):
    request = urllib.request.Request(
        "http://127.0.0.1:%d/json/list" % port,
        headers={"User-Agent": "Codex/1.0"},
    )
    with urllib.request.urlopen(request, timeout=5) as response:
        targets = json.loads(response.read().decode("utf-8", "ignore"))
    for target in targets:
        if target.get("type") == "page":
            return target.get("webSocketDebuggerUrl")
    raise RuntimeError("Edge page target not found")


def _cdp(ws, method, params, message_id):
    ws.send(json.dumps({"id": message_id, "method": method, "params": params or {}}))
    while True:
        message = json.loads(ws.recv())
        if message.get("id") == message_id:
            return message


def _set_cdp_cookies(ws, cookie_text):
    cookies = []
    for part in (cookie_text or "").split(";"):
        if "=" not in part:
            continue
        name, value = part.split("=", 1)
        name = name.strip()
        value = value.strip()
        if not name:
            continue
        cookies.append(
            {
                "name": name,
                "value": value,
                "domain": ".jd.com",
                "path": "/",
                "secure": True,
            }
        )
    if not cookies:
        return
    _cdp(ws, "Network.setCookies", {"cookies": cookies}, 1)


def _evaluate_cdp(ws, expression, await_promise=True, timeout=30):
    _cdp(ws, "Runtime.evaluate", {"expression": "1"}, 100)
    result = _cdp(
        ws,
        "Runtime.evaluate",
        {
            "expression": expression,
            "awaitPromise": await_promise,
            "returnByValue": True,
            "userGesture": True,
        },
        101,
    )
    return result


def _wait_page_ready(ws, timeout=30):
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            result = _evaluate_cdp(
                ws,
                "document.readyState === 'complete' && typeof window.SummerCryptico !== 'undefined'",
                False,
            )
            value = (
                result.get("result", {})
                .get("result", {})
                .get("value")
            )
            if value:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False


def call_jd_encrypted_headless(path, payload, cookie, user_id, app_code, key_text):
    edge_path = _find_edge_path()
    port = 10240 + secrets.randbelow(20000)
    profile = os.path.join(
        os.environ.get("TEMP", ROOT_DIR),
        "jd-unpack-cdp-" + secrets.token_hex(4),
    )
    process = subprocess.Popen(
        [
            edge_path,
            "--headless=new",
            "--disable-gpu",
            "--no-first-run",
            "--no-default-browser-check",
            "--remote-debugging-port=%d" % port,
            "--remote-allow-origins=*",
            "--user-data-dir=" + profile,
            "about:blank",
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    ws = None
    try:
        _wait_debugger(port)
        ws_url = _get_page_ws(port)
        ws = create_connection(ws_url, timeout=30)
        _cdp(ws, "Network.enable", {}, 2)
        _cdp(ws, "Page.enable", {}, 3)
        _set_cdp_cookies(ws, cookie)
        _cdp(
            ws,
            "Page.navigate",
            {"url": "https://digital-ins.jd.com/repair/business/pendingServiceList"},
            4,
        )
        if not _wait_page_ready(ws):
            raise RuntimeError("京东页面或 SummerCryptico 未加载完成")
        expression = (
            "(async () => {"
            " const keyResult = " + json.dumps(key_text) + ";"
            " const payload = " + json.dumps(payload, ensure_ascii=False) + ";"
            " return window.SummerCryptico.encryptData(keyResult, JSON.stringify(payload));"
            "})()"
        )
        result = _evaluate_cdp(ws, expression)
        exception = result.get("result", {}).get("exceptionDetails")
        if exception:
            raise RuntimeError(str(exception.get("exception", {}).get("description") or exception))
        encrypted_body = result.get("result", {}).get("result", {}).get("value")
        if not encrypted_body:
            raise RuntimeError("加密结果为空")
        return _post_encrypted_body(path, encrypted_body, cookie, user_id, app_code)
    finally:
        try:
            if ws:
                ws.close()
        except Exception:
            pass
        try:
            process.terminate()
        except Exception:
            pass
        try:
            import shutil

            shutil.rmtree(profile, ignore_errors=True)
        except Exception:
            pass


def _post_encrypted_body(path, encrypted_body, cookie, user_id, app_code):
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "identity",
        "Content-Type": "application/json",
        "Origin": "https://digital-ins.jd.com",
        "Referer": "https://digital-ins.jd.com/repair",
        "sysType": "1",
        "User-Agent": UA,
    }
    if cookie:
        headers["Cookie"] = cookie.replace("\r", "").replace("\n", "")
    if user_id:
        headers["userId"] = user_id.strip()
    if app_code:
        headers["appcode"] = app_code.strip()
    request = urllib.request.Request(
        JD_BASE + path,
        data=encrypted_body.encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            body = decode_body(response.read(), response.headers)
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = decode_body(error.read(), error.headers)
        try:
            return json.loads(body)
        except Exception:
            return {"success": False, "error": f"HTTP {error.code}: {body[:200]}"}
    except Exception as error:
        return {"success": False, "error": str(error)}


def call_jd_encrypted(path, payload, cookie, user_id, app_code):
    key_response = call_jd_get("/aks/getSMKey", cookie, user_id, app_code)
    sys.stdout.write(
        "bridge: getSMKey response=%s\n"
        % json.dumps(key_response, ensure_ascii=False)[:1000]
    )
    sys.stdout.flush()
    if not key_response.get("success"):
        return {
            "success": False,
            "error": str(
                key_response.get("showMsg")
                or key_response.get("msg")
                or key_response.get("error")
                or "获取加密密钥失败"
            ),
        }
    try:
        return call_jd_encrypted_headless(
            path, payload, cookie, user_id, app_code, key_response.get("result")
        )
    except Exception as error:
        sys.stdout.write(
            "bridge: headless encrypted error=%s\n" % str(error)
        )
        sys.stdout.flush()
    encrypted_body = jd_encrypt_data(key_response.get("result"), payload)
    sys.stdout.write(
        "bridge: encrypted body len=%d prefix=%s\n"
        % (len(encrypted_body), encrypted_body[:80])
    )
    sys.stdout.flush()
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "identity",
        "Content-Type": "application/json",
        "Origin": "https://digital-ins.jd.com",
        "Referer": "https://digital-ins.jd.com/repair",
        "sysType": "1",
        "User-Agent": UA,
    }
    if cookie:
        headers["Cookie"] = cookie.replace("\r", "").replace("\n", "")
    if user_id:
        headers["userId"] = user_id.strip()
    if app_code:
        headers["appcode"] = app_code.strip()
    request = urllib.request.Request(
        JD_BASE + path,
        data=encrypted_body.encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            body = decode_body(response.read(), response.headers)
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = decode_body(error.read(), error.headers)
        try:
            return json.loads(body)
        except Exception:
            return {"success": False, "error": f"HTTP {error.code}: {body[:200]}"}
    except Exception as error:
        return {"success": False, "error": str(error)}


def call_jd_service(path, payload, jdl_token, jdl_cookie=""):
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Origin": "https://jdservice.jdl.com",
        "Referer": "https://jdservice.jdl.com/spc/repair/servicebilllist",
        "User-Agent": UA,
        "login-type": "2",
        "X-Requested-With": "XMLHttpRequest",
    }
    token_value = (jdl_token or "").strip() or JDL_TOKEN
    cookie_value = (
        (jdl_cookie or "").strip().replace("\r", "").replace("\n", "")
        or JDL_COOKIE
    )
    if token_value:
        headers["X-Access-Token"] = token_value
    if cookie_value:
        headers["Cookie"] = cookie_value
    elif token_value:
        headers["Cookie"] = "pin=" + token_value
    request = urllib.request.Request(
        JD_SERVICE_BASE + path,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            body = decode_body(response.read(), response.headers)
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = decode_body(error.read(), error.headers)
        try:
            return json.loads(body)
        except Exception:
            return {"success": False, "error": f"HTTP {error.code}: {body[:200]}"}
    except Exception as error:
        return {"success": False, "error": str(error)}


def normalize_row(row):
    return {
        "performingOrderNo": find_key(row, "performingOrderNo") or find_key(row, "performingNo"),
        "serviceOrderNo": find_key(row, "serviceOrderNo"),
        "expressNo": find_key(row, "expressNo"),
        "facilitatorCode": find_key(row, "facilitatorCode"),
        "shopCode": find_key(row, "shopCode"),
        "serviceState": find_key(row, "serviceState"),
        "outerMainSkuThridCategory": find_key(row, "outerMainSkuThridCategory"),
        "outerMainSkuThirdCategory": find_key(row, "outerMainSkuThirdCategory"),
        "outerMainSkuName": find_key(row, "outerMainSkuName"),
        "outerSkuName": find_key(row, "outerSkuName"),
        "outerSku": find_key(row, "outerSku"),
    }


def format_performing_model(commit):
    leaf = find_key(commit, "leafPerformingModelName")
    code = find_key(commit, "performanceTypeEnum")
    if code in ("ADD_VALUE", "10") and leaf:
        return leaf
    return PERFORMANCE_MAP.get(code) or leaf or code or "未知状态"


def query_parts_barcode(
    merchant_order_no,
    jdl_token,
    jdl_cookie="",
    client_id="",
    service_bill_no="",
    afs_service_bill_no="",
):
    global LAST_BARCODE_ERROR
    LAST_BARCODE_ERROR = ""
    candidates = []
    if merchant_order_no:
        candidates.append(
            {"merchantOrderNo": str(merchant_order_no).strip()}
        )
    if service_bill_no:
        candidates.append({"serviceBillNo": str(service_bill_no).strip()})
    if afs_service_bill_no:
        candidates.append({"afsServiceBillNo": str(afs_service_bill_no).strip()})
    if not candidates:
        return ""
    token = (
        (jdl_token or "").strip()
        or CLIENT_JDL_TOKENS.get(client_id, "")
        or JDL_TOKEN
    )
    cookie = (
        (jdl_cookie or "").strip()
        or CLIENT_JDL_COOKIES.get(client_id, "")
        or JDL_COOKIE
    )
    for candidate in candidates:
        response = call_jd_service(
            "/spcapi/mcsServiceBill/page",
            {
                **candidate,
                "pageIndex": 1,
                "pageSize": 10,
                "serviceBillState": -1000,
                "createTimeBegin": None,
                "createTimeEnd": None,
            },
            token,
            cookie,
        )
        sys.stdout.write(
            "bridge: mcs page query=%r response=%s\n"
            % (candidate, json.dumps(response, ensure_ascii=False)[:2000])
        )
        sys.stdout.flush()
        if not isinstance(response, dict):
            LAST_BARCODE_ERROR = "京东物流返回格式异常"
            return ""
        if response.get("error") == "NotLogin" or response.get("success") is False:
            LAST_BARCODE_ERROR = str(
                response.get("error")
                or response.get("msg")
                or response.get("message")
                or "京东物流查询失败"
            )
            return ""
        data = response.get("data") or {}
        item_list = data.get("itemList") or []
        if item_list and isinstance(item_list, list):
            first = item_list[0]
            if isinstance(first, dict) and first.get("partCode"):
                return str(first["partCode"]).strip()
    LAST_BARCODE_ERROR = "京东物流未返回备件条码"
    return ""


def query_repair(
    express_no,
    cookie,
    user_id,
    app_code,
    shop_code,
    jdl_token,
    jdl_cookie="",
    client_id="",
):
    if not express_no or not str(express_no).strip():
        return {"ok": False, "error": "快递单号不能为空"}

    client_config = CLIENT_CONFIGS.get(client_id) or {}
    cookie = str(cookie or "").strip() or client_config.get("cookie", "") or DIGITAL_CONFIG.get("cookie", "")
    user_id = str(user_id or "").strip() or client_config.get("userId", "") or DIGITAL_CONFIG.get("userId", "")
    app_code = str(app_code or "").strip() or client_config.get("appCode", "") or DIGITAL_CONFIG.get("appCode", "")
    shop_code = str(shop_code or "").strip() or client_config.get("shopCode", "") or DIGITAL_CONFIG.get("shopCode", "")
    if not user_id:
        user_id = extract_cookie_value(cookie, "pin")
    if not app_code:
        app_code = extract_cookie_value(cookie, "systemCode")
    if not shop_code:
        shop_code = extract_cookie_value(cookie, "shopCode")
    facilitator_code = ""
    shop_name = ""
    shops = call_jd(
        "/serviceOrder/queryBindShopInfo",
        {},
        cookie,
        user_id,
        app_code,
    )
    if shops.get("success") and shops.get("values"):
        first_shop = shops["values"][0]
        shop_code = str(first_shop.get("code") or "").strip()
        facilitator_code = str(first_shop.get("facilitatorCode") or "").strip()
        shop_name = str(
            first_shop.get("shopName")
            or first_shop.get("name")
            or ""
        ).strip()

    code_value = str(express_no).strip().upper()
    performing_style = code_value.isdigit() or (
        code_value.startswith("JD") and code_value[2:].isdigit()
    )
    if code_value.startswith("JDX") or not performing_style:
        query_fields = [
            "expressNo",
            "performingId",
            "serviceOrderNo",
            "orderId",
        ]
    else:
        query_fields = [
            "performingId",
            "serviceOrderNo",
            "orderId",
        ]
    rows = []
    list_response = None
    for field in query_fields:
        query = {field: code_value}
        if shop_code:
            query["shopCode"] = shop_code.strip()
        list_response = call_jd(
            "/serviceOrder/queryFacilitatorOrderByStateAndCode",
            {
                "pageIndex": 1,
                "pageSize": 10,
                "query": query,
            },
            cookie,
            user_id,
            app_code,
        )
        if list_response.get("success"):
            rows = list_response.get("values") or list_response.get("value") or []
            if isinstance(rows, dict):
                rows = [rows]
            if rows:
                break
    if not list_response:
        return {"ok": False, "error": "查询接口未返回结果"}
    if not rows:
        if not list_response.get("success"):
            message = (
                list_response.get("msg")
                or list_response.get("error")
                or list_response.get("showMsg")
                or "查询失败"
            )
            return {"ok": False, "error": str(message)}
        return {"ok": True, "found": False}

    row = rows[0]
    row_info = normalize_row(row)
    detail_payload = {
        "serviceOrderNo": row_info["serviceOrderNo"],
        "facilitatorCode": row_info["facilitatorCode"] or facilitator_code,
        "shopCode": row_info["shopCode"] or shop_code,
        "performingOrderState": row_info["serviceState"] or find_key(row, "performingOrderState"),
    }
    detail_response = call_jd(
        "/serviceOrder/serviceOrderDetail",
        detail_payload,
        cookie,
        user_id,
        app_code,
    )
    if not detail_response.get("success"):
        message = (
            detail_response.get("msg")
            or detail_response.get("error")
            or detail_response.get("showMsg")
            or "详情查询失败"
        )
        return {"ok": False, "error": str(message), "row": row}

    detail = detail_response.get("value") or {}
    commit = detail.get("repairCommitInfoDto") or {}
    receive = detail.get("receiveMaintainOrderInfoDto") or {}
    express_list = find_key(receive, "expressInfoDtoList")
    express_no = row_info["expressNo"] or find_key(receive, "expressNo")
    if not express_no and isinstance(express_list, list) and express_list:
        first = express_list[0]
        express_no = first.get("expressNo") if isinstance(first, dict) else first
    if not express_no:
        express_no = str(express_no).strip().upper()
    performing_order_no = (
        find_key(commit, "performingOrderNo")
        or find_key(commit, "performingNo")
        or find_key(receive, "performingOrderNo")
        or find_key(receive, "performingNo")
        or find_key(detail, "performingOrderNo")
        or find_key(detail, "performingNo")
        or row_info["performingOrderNo"]
        or row_info["serviceOrderNo"]
    )
    merchant_order_no = (
        find_key(commit, "merchantOrderNo")
        or find_key(receive, "merchantOrderNo")
        or find_key(detail, "merchantOrderNo")
    )
    service_bill_no = (
        find_key(commit, "serviceBillNo")
        or find_key(receive, "serviceBillNo")
        or find_key(detail, "serviceBillNo")
    )
    afs_service_bill_no = (
        find_key(commit, "afsServiceBillNo")
        or find_key(receive, "afsServiceBillNo")
        or find_key(detail, "afsServiceBillNo")
    )
    part_barcode = query_parts_barcode(
        merchant_order_no or performing_order_no,
        jdl_token,
        jdl_cookie,
        client_id,
        service_bill_no,
        afs_service_bill_no,
    )

    return {
        "ok": True,
        "found": True,
        "performingOrderNo": performing_order_no,
        "merchantOrderNo": merchant_order_no,
        "serviceBillNo": service_bill_no,
        "afsServiceBillNo": afs_service_bill_no,
        "serviceOrderNo": row_info["serviceOrderNo"] or find_key(commit, "serviceOrderNo"),
        "category": (
            find_key(commit, "outerMainSkuThridCategory")
            or find_key(commit, "outerMainSkuThirdCategory")
            or row_info["outerMainSkuThridCategory"]
            or row_info["outerMainSkuThirdCategory"]
        ),
        "expressNo": express_no,
        "productName": (
            find_key(commit, "outerMainSkuName")
            or find_key(commit, "outerSkuName")
            or row_info["outerMainSkuName"]
            or row_info["outerSkuName"]
        ),
        "sku": find_key(commit, "outerSku") or row_info["outerSku"],
        "brand": find_key(commit, "outerMainSkuBrand"),
        "model": find_key(commit, "mainSkuModel"),
        "performingModel": format_performing_model(commit),
        "shopName": shop_name,
        "partBarcode": part_barcode,
        "row": row,
        "detail": detail,
    }


def auto_start_and_sync(
    express_no,
    cookie,
    user_id,
    app_code,
    shop_code,
    jdl_token,
    jdl_cookie="",
    client_id="",
):
    steps = []
    base = query_repair(
        express_no,
        cookie,
        user_id,
        app_code,
        shop_code,
        jdl_token,
        jdl_cookie,
        client_id,
    )
    steps.append(
        {
            "name": "query",
            "ok": bool(base.get("ok")),
            "found": bool(base.get("found")),
            "error": base.get("error"),
        }
    )
    if not base.get("ok") or not base.get("found"):
        return {**base, "steps": steps}

    detail = base.get("detail") or {}
    commit = detail.get("repairCommitInfoDto") or {}
    service_id = str(
        find_key(commit, "facilitatorCode") or find_key(detail, "facilitatorCode") or ""
    ).strip()
    shop_id = str(
        find_key(commit, "shopCode") or find_key(detail, "shopCode") or ""
    ).strip()
    service_order_no = str(
        base.get("serviceOrderNo")
        or find_key(commit, "serviceOrderNo")
        or find_key(detail, "serviceOrderNo")
        or ""
    ).strip()
    service_type = str(
        find_key(commit, "serviceTypeEnum")
        or find_key(detail, "serviceTypeEnum")
        or ""
    ).strip()
    service_state = str(
        find_key(commit, "serviceState")
        or find_key(detail, "serviceState")
        or ""
    ).strip()
    wait_receive_states = {
        "EXPRESS_INFO_FINISH",
        "WAIT_VISIT",
        "WAIT_ARRIVE_STORE",
        "WAIT_RECEIVE",
        "WAIT_RECEIVED",
        "TO_BE_RECEIVED",
        "RECEIVE",
    }
    already_started = bool(service_state) and service_state not in wait_receive_states

    order_info = {}
    for key in (
        "storePhone",
        "engineerIndex",
        "engineerJdPin",
        "engineerName",
        "engineerPhone",
        "maintainType",
        "modifyCallCause",
        "subscribeTime",
        "subscribeTimeType",
        "returnAddress",
        "warehouseCode",
        "needReturnGoods",
    ):
        value = find_key(commit, key) or find_key(detail, key)
        if value not in (None, ""):
            order_info[key] = str(value).strip() if not isinstance(value, bool) else value

    receive = detail.get("receiveMaintainOrderInfoDto") or {}
    if not order_info.get("storePhone"):
        shop_tel = find_key(receive, "shopTel") or find_key(detail, "shopTel")
        if shop_tel not in (None, ""):
            order_info["storePhone"] = str(shop_tel).strip()
    for key in (
        "engineerIndex",
        "engineerJdPin",
        "engineerName",
        "engineerPhone",
        "maintainType",
        "modifyCallCause",
        "subscribeTime",
        "subscribeTimeType",
    ):
        if not order_info.get(key):
            value = find_key(receive, key) or find_key(detail, key)
            if value not in (None, ""):
                order_info[key] = str(value).strip()

    if not order_info.get("warehouseCode"):
        warehouse_response = call_jd(
            "/warehouse/queryByFacilitatorAndShopCode",
            {"shopCode": shop_id, "busStatus": "1"},
            cookie,
            user_id,
            app_code,
        )
        sys.stdout.write(
            "bridge: warehouse query response=%s\n"
            % json.dumps(warehouse_response, ensure_ascii=False)[:1000]
        )
        sys.stdout.flush()
        warehouse_values = warehouse_response.get("values") or []
        if warehouse_values and isinstance(warehouse_values, list):
            first_warehouse = warehouse_values[0]
            if isinstance(first_warehouse, dict):
                order_info["warehouseCode"] = str(
                    first_warehouse.get("warehouseCode") or ""
                ).strip()

    if not order_info.get("returnAddress") and order_info.get("warehouseCode"):
        warehouse_detail_response = call_jd_get(
            "/warehouse/queryByWarehouseCode/" + order_info["warehouseCode"],
            cookie,
            user_id,
            app_code,
        )
        sys.stdout.write(
            "bridge: warehouse detail response=%s\n"
            % json.dumps(warehouse_detail_response, ensure_ascii=False)[:1000]
        )
        sys.stdout.flush()
        warehouse_detail = warehouse_detail_response.get("value") or {}
        if isinstance(warehouse_detail, dict):
            address_parts = [
                warehouse_detail.get("provinceName"),
                warehouse_detail.get("cityName"),
                warehouse_detail.get("countyName"),
                warehouse_detail.get("townName"),
                warehouse_detail.get("addressDetail"),
            ]
            address_parts = [str(part).strip() for part in address_parts if part]
            if address_parts:
                order_info["returnAddress"] = "/".join(address_parts)

    if not service_order_no:
        return {
            **base,
            "steps": steps,
            "startSuccess": False,
            "error": "未获取到服务单号，无法自动开始接机",
        }

    start_response = {}
    start_success = True
    if already_started:
        start_response = {"success": True, "showMsg": "该单已接机，跳过开始接机"}
    else:
        start_payload = {
            "serviceId": service_id,
            "shopId": shop_id,
            "serviceOrderNo": service_order_no,
            "serviceType": service_type,
            **order_info,
        }
        start_response = call_jd_encrypted(
            "/serviceOrder/facilitatorAcceptExpress",
            start_payload,
            cookie,
            user_id,
            app_code,
        )
        sys.stdout.write(
            "bridge: auto-start payload=%s response=%s\n"
            % (
                json.dumps(start_payload, ensure_ascii=False),
                json.dumps(start_response, ensure_ascii=False)[:1200],
            )
        )
        sys.stdout.flush()
        start_success = bool(start_response.get("success"))
        start_error_text = str(
            start_response.get("showMsg")
            or start_response.get("msg")
            or start_response.get("error")
            or ""
        )
        if "HTTP 400" in start_error_text or "HTTP 4" in start_error_text:
            start_response["encryptionRequired"] = True
            start_response["showMsg"] = "需要在京东维修页面执行自动接机命令"
        if not start_success and any(
            keyword in start_error_text
            for keyword in ("已接机", "已收货", "已确认接机", "重复操作", "已开始")
        ):
            start_success = True
            start_response["showMsg"] = start_error_text
    steps.append(
        {
            "name": "start",
            "ok": start_success,
            "skipped": already_started,
            "error": start_response.get("showMsg")
            or start_response.get("msg")
            or start_response.get("error"),
            "response": start_response,
        }
    )

    part_barcode = ""
    performing_order_no = base.get("performingOrderNo") or ""
    merchant_order_no = base.get("merchantOrderNo") or ""
    service_bill_no = base.get("serviceBillNo") or ""
    afs_service_bill_no = base.get("afsServiceBillNo") or ""
    barcode_attempts = 0
    max_barcode_attempts = 1
    while not part_barcode and barcode_attempts < max_barcode_attempts:
        barcode_attempts += 1
        if performing_order_no or merchant_order_no:
            part_barcode = query_parts_barcode(
                merchant_order_no or performing_order_no,
                jdl_token,
                jdl_cookie,
                client_id,
                service_bill_no,
                afs_service_bill_no,
            )
        sys.stdout.write(
            "bridge: barcode retry %d/%d found=%s partCode=%r error=%r\n"
            % (
                barcode_attempts,
                max_barcode_attempts,
                bool(part_barcode),
                part_barcode,
                LAST_BARCODE_ERROR,
            )
        )
        sys.stdout.flush()
    steps.append(
        {
            "name": "barcode",
            "ok": bool(part_barcode),
            "partBarcode": part_barcode,
            "retries": barcode_attempts,
            "error": LAST_BARCODE_ERROR,
        }
    )

    result = {**base, "steps": steps, "startSuccess": start_success}
    if start_response.get("encryptionRequired"):
        result["encryptionRequired"] = True
    if part_barcode:
        result["partBarcode"] = part_barcode
    if not start_success:
        if start_response.get("encryptionRequired"):
            result["error"] = "请在京东维修页面执行自动接机命令，再回来查询"
        else:
            result["error"] = "自动开始接机失败：" + str(
                steps[-2].get("error") or "请到京东维修页面手动操作"
            )
    return result


def desktop_dir():
    home = os.path.expanduser("~")
    for candidate in (
        os.path.join(home, "Desktop"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(home, "OneDrive", "桌面"),
    ):
        if os.path.isdir(candidate):
            return candidate
    return home


def export_categories_xlsx(rows, account="admin"):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "品类看板"
    headers = ["品类", "今日数量", "累计数量", "今日占比"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="E1251B")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    normalized_rows = []
    for item in rows or []:
        category = str(item.get("category") or "未分类")
        today_count = int(item.get("todayCount") or 0)
        total_count = int(item.get("totalCount") or item.get("count") or 0)
        normalized_rows.append([category, today_count, total_count])
    total_all = sum(row[2] for row in normalized_rows) or 1
    for category, today_count, total_count in normalized_rows:
        percent = round((total_count / total_all) * 100, 1)
        ws.append([category, today_count, total_count, percent])

    for col, width in zip("ABCD", [28, 12, 12, 12]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    filename = "品类看板_%s_%s.xlsx" % (
        account,
        time.strftime("%Y%m%d_%H%M%S"),
    )
    path = os.path.join(desktop_dir(), filename)
    wb.save(path)
    return path


class BridgeHandler(BaseHTTPRequestHandler):
    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Cache-Control", "no-store")

    def _send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self):
        global JDL_TOKEN, JDL_COOKIE
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/health":
            self._send_json(200, {"ok": True, "message": "JD repair bridge is running"})
            return
        if parsed.path == "/api/repair/set-digital-cookie":
            query = urllib.parse.parse_qs(parsed.query)
            data_text = (query.get("data") or [""])[0]
            try:
                payload = json.loads(data_text)
            except Exception:
                self._send_json(400, {"ok": False, "error": "data 不是合法 JSON"})
                return
            client_id = str(payload.get("clientId", "") or "").strip()
            cookie_text = str(payload.get("cookie", "") or "").strip()
            config_updates = {
                "cookie": cookie_text,
                "userId": str(payload.get("userId", "") or extract_cookie_value(cookie_text, "pin") or "").strip(),
                "appCode": str(payload.get("appCode", "") or extract_cookie_value(cookie_text, "systemCode") or "").strip(),
                "shopCode": str(payload.get("shopCode", "") or extract_cookie_value(cookie_text, "shopCode") or "").strip(),
            }
            jdl_token = str(payload.get("jdlToken", "") or "").strip()
            jdl_cookie = str(payload.get("jdlCookie", "") or "").strip()
            if jdl_token or jdl_cookie:
                if jdl_token:
                    JDL_TOKEN = jdl_token
                if jdl_cookie:
                    JDL_COOKIE = jdl_cookie
                if client_id:
                    if jdl_token:
                        CLIENT_JDL_TOKENS[client_id] = jdl_token
                    if jdl_cookie:
                        CLIENT_JDL_COOKIES[client_id] = jdl_cookie
                save_jdl_token()
            DIGITAL_CONFIG.update(config_updates)
            if client_id:
                CLIENT_CONFIGS[client_id] = {
                    **(CLIENT_CONFIGS.get(client_id) or {}),
                    **config_updates,
                }
            save_digital_config()
            self._send_json(200, {"ok": True, "message": "OK"})
            return
        if parsed.path == "/api/repair/set-jdl-token":
            query = urllib.parse.parse_qs(parsed.query)
            data_text = (query.get("data") or [""])[0]
            try:
                payload = json.loads(data_text)
            except Exception:
                self._send_json(400, {"ok": False, "error": "data 不是合法 JSON"})
                return
            client_id = str(payload.get("clientId", "") or "").strip()
            token = str(payload.get("jdlToken", "") or "").strip()
            cookie = str(payload.get("jdlCookie", "") or "").strip()
            if token or cookie:
                JDL_TOKEN = token
                JDL_COOKIE = cookie
                if client_id:
                    CLIENT_JDL_TOKENS[client_id] = token
                    CLIENT_JDL_COOKIES[client_id] = cookie
                save_jdl_token()
            self._send_json(
                200,
                {
                    "ok": True,
                    "message": "OK",
                    "configured": bool(JDL_TOKEN or JDL_COOKIE),
                },
            )
            return
        if parsed.path == "/api/repair/config":
            token_configured = bool(
                JDL_TOKEN
                or any(CLIENT_JDL_TOKENS.values())
                or JDL_COOKIE
                or any(CLIENT_JDL_COOKIES.values())
            )
            self._send_json(
                200,
                {
                    "ok": True,
                    "digitalConfigured": bool(DIGITAL_CONFIG.get("cookie")),
                    "jdlConfigured": token_configured,
                    "message": "京东维修/展翅登录状态",
                },
            )
            return
        if parsed.path == "/api/repair/latest":
            query = urllib.parse.parse_qs(parsed.query)
            tracking = (query.get("tracking") or [""])[0].strip().upper()
            result = LATEST_RESULTS.get(tracking)
            if result:
                self._send_json(200, result)
            else:
                self._send_json(200, {"ok": True, "found": False, "tracking": tracking})
            return

        if parsed.path == "/api/print/find":
            printer_path = find_barcode_printer()
            self._send_json(
                200,
                {
                    "ok": bool(printer_path),
                    "path": printer_path,
                    "error": None if printer_path else "未在本机找到 BarPrinter.exe",
                },
            )
            return

        if parsed.path == "/api/state":
            query = urllib.parse.parse_qs(parsed.query)
            account = (query.get("account") or ["admin"])[0]
            self._send_json(
                200,
                {
                    "ok": True,
                    "account": account,
                    "state": SHARED_STATES.get(account) or {
                        "parcels": [],
                        "anomalies": [],
                    },
                },
            )
            return

        relative = parsed.path.lstrip("/") or "index.html"
        file_path = os.path.realpath(os.path.join(ROOT_DIR, relative))
        if not file_path.startswith(ROOT_DIR) or not os.path.isfile(file_path):
            self._send_json(404, {"ok": False, "error": "Not found"})
            return
        content_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
        with open(file_path, "rb") as handle:
            body = handle.read()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        global JDL_TOKEN, JDL_COOKIE
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/print":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            order_no = str(payload.get("orderNo", "") or "").strip()
            part_code = str(payload.get("partCode", "") or "").strip()
            if not order_no:
                self._send_json(400, {"ok": False, "error": "缺少履约单号"})
                return
            command = [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                PRINT_SCRIPT,
                "-OrderNo",
                order_no,
            ]
            bar_printer_path = str(payload.get("barPrinterPath", "") or "").strip()
            if bar_printer_path:
                command += ["-BarPrinterPath", bar_printer_path]
            if payload.get("dryRun"):
                command += ["-DryRun"]
            try:
                result = subprocess.run(
                    command,
                    capture_output=True,
                    text=True,
                    timeout=60,
                )
            except subprocess.TimeoutExpired:
                self._send_json(504, {"ok": False, "error": "条码打印超时"})
                return
            except Exception as error:
                self._send_json(500, {"ok": False, "error": str(error)})
                return
            output = (result.stdout or "") + (result.stderr or "")
            self._send_json(
                200,
                {
                    "ok": result.returncode == 0,
                    "output": output,
                    "error": None if result.returncode == 0 else output,
                },
            )
            return

        if parsed.path == "/api/state":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            account = str(payload.get("account") or "admin").strip()
            incoming = payload.get("state") or {}
            stored = SHARED_STATES.get(account) or {
                "parcels": [],
                "anomalies": [],
            }
            if payload.get("replace"):
                merged = {
                    "parcels": incoming.get("parcels") or [],
                    "anomalies": incoming.get("anomalies") or [],
                    "clearedAt": payload.get("clearedAt")
                    or stored.get("clearedAt")
                    or "",
                }
            else:
                merged = merge_shared_state(stored, incoming)
                if stored.get("clearedAt"):
                    merged["clearedAt"] = stored["clearedAt"]
            SHARED_STATES[account] = merged
            save_shared_states()
            self._send_json(
                200,
                {
                    "ok": True,
                    "account": account,
                    "state": merged,
                },
            )
            return

        if parsed.path == "/api/export-categories":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            account = str(payload.get("account") or "admin").strip()
            rows = payload.get("rows") or []
            try:
                path = export_categories_xlsx(rows, account)
                self._send_json(
                    200,
                    {"ok": True, "path": path, "message": "已导出到桌面"},
                )
            except Exception as error:
                self._send_json(500, {"ok": False, "error": str(error)})
            return

        if parsed.path == "/api/repair/set-digital-cookie":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            client_id = str(payload.get("clientId", "") or "").strip()
            config_updates = {
                "cookie": str(payload.get("cookie", "") or "").strip(),
                "userId": str(payload.get("userId", "") or "").strip(),
                "appCode": str(payload.get("appCode", "") or "").strip(),
                "shopCode": str(payload.get("shopCode", "") or "").strip(),
            }
            jdl_token = str(payload.get("jdlToken", "") or "").strip()
            jdl_cookie = str(payload.get("jdlCookie", "") or "").strip()
            if jdl_token or jdl_cookie:
                if jdl_token:
                    JDL_TOKEN = jdl_token
                if jdl_cookie:
                    JDL_COOKIE = jdl_cookie
                if client_id:
                    if jdl_token:
                        CLIENT_JDL_TOKENS[client_id] = jdl_token
                    if jdl_cookie:
                        CLIENT_JDL_COOKIES[client_id] = jdl_cookie
                save_jdl_token()
            DIGITAL_CONFIG.update(config_updates)
            if client_id:
                CLIENT_CONFIGS[client_id] = {
                    **(CLIENT_CONFIGS.get(client_id) or {}),
                    **config_updates,
                }
            save_digital_config()
            self._send_json(
                200,
                {
                    "ok": True,
                    "configured": bool(DIGITAL_CONFIG["cookie"]),
                    "message": "京东维修登录信息已保存到本地桥接服务",
                },
            )
            return

        if parsed.path == "/api/repair/set-jdl-token":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            client_id = str(payload.get("clientId", "") or "").strip()
            token = str(payload.get("jdlToken", "") or "").strip()
            cookie = str(payload.get("jdlCookie", "") or "").strip()
            JDL_TOKEN = token
            JDL_COOKIE = cookie
            save_jdl_token()
            if client_id:
                CLIENT_JDL_TOKENS[client_id] = token
                CLIENT_JDL_COOKIES[client_id] = cookie
            self._send_json(
                200,
                {
                    "ok": True,
                    "configured": bool(JDL_TOKEN or JDL_COOKIE),
                    "message": "京东物流 Token 已保存到本地桥接服务",
                },
            )
            return

        if parsed.path == "/api/repair/import":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                raw_body = self.rfile.read(length).decode("utf-8")
                content_type = self.headers.get("Content-Type", "")
                if "application/x-www-form-urlencoded" in content_type:
                    form = urllib.parse.parse_qs(raw_body)
                    payload = json.loads((form.get("data") or [""])[0])
                else:
                    payload = json.loads(raw_body)
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            tracking = str(payload.get("tracking", "")).strip().upper()
            result = payload.get("result")
            if not tracking or not isinstance(result, dict):
                self._send_json(400, {"ok": False, "error": "缺少 tracking 或 result"})
                return
            jdl_token = str(payload.get("jdlToken", "") or "").strip()
            jdl_cookie = str(payload.get("jdlCookie", "") or "").strip()
            if jdl_token or jdl_cookie:
                JDL_TOKEN = jdl_token
                JDL_COOKIE = jdl_cookie
                client_id = str(payload.get("clientId", "") or "").strip()
                if client_id:
                    CLIENT_JDL_TOKENS[client_id] = jdl_token
                    CLIENT_JDL_COOKIES[client_id] = jdl_cookie
                save_jdl_token()
            LATEST_RESULTS[tracking] = result
            self._send_json(200, {"ok": True, "tracking": tracking})
            return

        if parsed.path == "/api/repair/part-barcode":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            merchant_order_no = payload.get("merchantOrderNo", "")
            jdl_token = payload.get("jdlToken", "")
            jdl_cookie = payload.get("jdlCookie", "")
            client_id = payload.get("clientId", "")
            part_barcode = query_parts_barcode(
                merchant_order_no,
                jdl_token,
                jdl_cookie,
                client_id,
                payload.get("serviceBillNo", ""),
                payload.get("afsServiceBillNo", ""),
            )
            sys.stdout.write(
                "bridge: part-barcode order=%r found=%s partCode=%r error=%r\n"
                % (
                    merchant_order_no,
                    bool(part_barcode),
                    part_barcode,
                    LAST_BARCODE_ERROR,
                )
            )
            sys.stdout.flush()
            self._send_json(
                200,
                {
                    "ok": True,
                    "found": bool(part_barcode),
                    "partBarcode": part_barcode,
                    "merchantOrderNo": merchant_order_no,
                    "error": LAST_BARCODE_ERROR
                    or (None if part_barcode else "未查询到备件条码"),
                },
            )
            return

        if parsed.path == "/api/repair/auto-start":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
                return
            result = auto_start_and_sync(
                payload.get("expressNo", ""),
                payload.get("cookie", ""),
                payload.get("userId", ""),
                payload.get("appCode", ""),
                payload.get("shopCode", ""),
                payload.get("jdlToken", ""),
                payload.get("jdlCookie", ""),
                payload.get("clientId", ""),
            )
            self._send_json(200, result)
            return

        if parsed.path != "/api/repair/query":
            self._send_json(404, {"ok": False, "error": "Not found"})
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except Exception:
            self._send_json(400, {"ok": False, "error": "请求体不是合法 JSON"})
            return

        express_no = payload.get("expressNo", "")
        cookie = payload.get("cookie", "")
        user_id = payload.get("userId", "")
        app_code = payload.get("appCode", "")
        shop_code = payload.get("shopCode", "")
        jdl_token = payload.get("jdlToken", "")
        jdl_cookie = payload.get("jdlCookie", "")
        client_id = payload.get("clientId", "")
        result = query_repair(
            express_no,
            cookie,
            user_id,
            app_code,
            shop_code,
            jdl_token,
            jdl_cookie,
            client_id,
        )
        sys.stdout.write(
            "bridge: query %r ok=%s found=%s performing=%r service=%r merchant=%r error=%r\n"
            % (
                express_no,
                result.get("ok"),
                result.get("found"),
                result.get("performingOrderNo"),
                result.get("serviceOrderNo"),
                result.get("merchantOrderNo"),
                result.get("error"),
            )
        )
        sys.stdout.flush()
        self._send_json(200, result)

    def log_message(self, format, *args):
        sys.stdout.write("bridge: " + format % args + "\n")
        sys.stdout.flush()


def main():
    start_port = int(sys.argv[1]) if len(sys.argv) > 1 else 9090
    server = None
    port = None
    for candidate in range(start_port, start_port + 20):
        try:
            server = ThreadingHTTPServer(("0.0.0.0", candidate), BridgeHandler)
            port = candidate
            break
        except OSError:
            continue
    if server is None or port is None:
        print("JD repair bridge: no free port", flush=True)
        sys.exit(1)
    try:
        hostname = socket.gethostname()
        local_ips = list(dict.fromkeys(socket.gethostbyname_ex(hostname)[2]))
    except Exception:
        local_ips = ["127.0.0.1"]
    print(f"JD repair bridge ready at http://127.0.0.1:{port}", flush=True)
    for ip in local_ips:
        print(f"LAN: http://{ip}:{port}", flush=True)
    load_digital_config()
    load_jdl_token()
    load_shared_states()
    server.serve_forever()


if __name__ == "__main__":
    main()
